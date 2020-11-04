#!/usr/bin/env python
# coding: utf-8

# In[63]:


# Ranging #
from pypozyx import *
import numpy as np
from pypozyx.definitions.registers import *
import sys
import time
import serial 
from openpyxl import Workbook
write_wb = Workbook()
write_ws = write_wb.active

def cal(cira,th):
    idx = 0
    cir_cal = 0
    l = len(cira)
    mag = abs(cira[:-36])
    for i in range(1,993):
        if(mag[i]>th):
            idx = i
            break
    tmp = mag[idx]-500
    mag_idx = 0
    for i in range(idx+1,993):
        if(mag[i] > tmp and mag[i] > th):
            mag_idx = i
            tmp = mag[i]
    tmp_idx = mag_idx - idx
    if (mag_idx - idx < 40):
         tmp_idx = 0
    cir_cal= tmp_idx*64*3*pow(10,8)/(499.2*pow(10,6))
    print(mag_idx-idx)
    return cir_cal
i = 1    
port = '/dev/cu.usbmodem3998326833361'
p = PozyxSerial(port)

remote_id = 0x675a          # the network ID of the remote device
remote = False               # whether to use the given remote device for ranging
if not remote:
    remote_id = None
th = 1000
while True:
    ser = serial.Serial(port=port,baudrate=28800, timeout = 1)
    #print ser

    destination_id_1 = 0x6758      # network ID of the ranging destination
    range_step_mm = 1000         # distance that separates the amount of LEDs lighting up.
    device_range_1 = DeviceRange()
    status_1 = p.doRanging(destination_id_1, device_range_1, remote_id)
    zd_1= device_range_1[1]
    print(device_range_1[1])
    if status_1:
        list_offset = range(0, 1015, 49)
        data_length = 49
        sys.stdout.flush()
        ser.nonblocking()
        ser.flushInput()  #flush input buffer, discarding all its contents
        ser.flushOutput()
        ser.nonblocking()
        cir_buffer = Buffer([0] * 98 * len(list_offset), size=2, signed=1)
        status_cir = p.getDeviceCir(list(list_offset), data_length, cir_buffer, remote_id)
        if status_cir:
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                #get real and imaginarypart of the cir buffer
                real = np.array(cir_buffer.data[0::2])
                imag = np.array(cir_buffer.data[1::2])
                # create an image of the CIR
                cira = real + 1j*imag
                #print(device_range)
                #mat[i][j] = dist
                #print(dist)
                #That plots the CIR contains in the buffer.
                #It still requires post-procesing to
                #re-align delay and received power level.
                #plt.plot((abs(cira[:-36])))
                dist_1 = device_range_1[1]-cal(cira,th)
                print(dist_1)
                #plt.show()
            except:
                print('Buffer Error on 1')
        else:
            print('error in getting cir')
    else:
        print('Ranging failed on 1')
        zd_1= device_range_1[1]
    sys.stdout.flush()
    ser.flushInput()  #flush input buffer, discarding all its contents
    ser.flushOutput()
    destination_id_2 = 0x6714      # network ID of the ranging destination
    range_step_mm = 1000         # distance that separates the amount of LEDs lighting up.
    device_range_2 = DeviceRange()
    status_2 = p.doRanging(destination_id_2, device_range_2, remote_id)
    zd_2= device_range_2[1]
    print(device_range_2[1])
    if status_2:
        list_offset = range(0, 1015, 49)
        data_length = 49
        sys.stdout.flush()
        ser.nonblocking()
        ser.flushInput()  #flush input buffer, discarding all its contents
        ser.flushOutput()
        cir_buffer = Buffer([0] * 98 * len(list_offset), size=2, signed=1)
        status_cir = p.getDeviceCir(list(list_offset), data_length, cir_buffer, remote_id)
        if status_cir:
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                #get real and imaginarypart of the cir buffer
                real = np.array(cir_buffer.data[0::2])
                imag = np.array(cir_buffer.data[1::2])
                # create an image of the CIR
                cira = real + 1j*imag
                #print(device_range)
                #mat[i][j] = dist
                #print(dist)
                #That plots the CIR contains in the buffer.
                #It still requires post-procesing to
                #re-align delay and received power level.
                #plt.plot((abs(cira[:-36])))
                dist_2 = device_range_2[1]-cal(cira,th)
                print(dist_2)
                #plt.show()
            except:
                print('Buffer error on 2')
        else:
            print('error in getting cir')
    else:
        print('Ranging failed on 2')
        zd_2= device_range_2[1]

    #destination_id_3 = 0x6a32      # network ID of the ranging destination
    destination_id_3 = 0x6a5e
    range_step_mm = 1000         # distance that separates the amount of LEDs lighting up.
    device_range_3 = DeviceRange()
    status_3 = p.doRanging(destination_id_3, device_range_3, remote_id)
    zd_3= device_range_3[1]
    print(device_range_3[1])
    if status_3:
        list_offset = range(0, 1015, 49)
        data_length = 49
        sys.stdout.flush()
        ser.nonblocking()
        ser.flushInput()  #flush input buffer, discarding all its contents
        ser.flushOutput()
        cir_buffer = Buffer([0] * 98 * len(list_offset), size=2, signed=1)
        status_cir = p.getDeviceCir(list(list_offset), data_length, cir_buffer, remote_id)
        if status_cir:
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                    #get real and imaginarypart of the cir buffer
                real = np.array(cir_buffer.data[0::2])
                imag = np.array(cir_buffer.data[1::2])
                    # create an image of the CIR
                cira = real + 1j*imag
                #print(device_range)
                #mat[i][j] = dist
                #print(dist)
                #That plots the CIR contains in the buffer.
                #It still requires post-procesing to
                #re-align delay and received power level.
                #plt.plot((abs(cira[:-36])))
                dist_3 = device_range_3[1]-cal(cira,th)
                print(dist_3)
                #plt.show()
            except:
                print('Buffer Error on 3')
        else:
            print('error in getting cir')
    else:
        print('Ranging failed on 3')
        zd_3= device_range_3[1]

    ser.flushInput()  #flush input buffer, discarding all its contents
    ser.flushOutput()
    destination_id_4 = 0x6e6e      # network ID of the ranging destination
    range_step_mm = 1000         # distance that separates the amount of LEDs lighting up.
    device_range_4 = DeviceRange()
    status_4 = p.doRanging(destination_id_4, device_range_4, remote_id)
    zd_4= device_range_4[1]
    print(device_range_4[1])
    if status_4:
        list_offset = range(0, 1015, 49)
        data_length = 49
        sys.stdout.flush()
        ser.nonblocking()
        ser.flushInput()  #flush input buffer, discarding all its contents
        ser.flushOutput()
        cir_buffer = Buffer([0] * 98 * len(list_offset), size=2, signed=1)
        status_cir = p.getDeviceCir(list(list_offset), data_length, cir_buffer, remote_id)
        if status_cir:
            try:
                import matplotlib.pyplot as plt
                import numpy as np
                #get real and imaginarypart of the cir buffer
                real = np.array(cir_buffer.data[0::2])
                imag = np.array(cir_buffer.data[1::2])
                # create an image of the CIR
                cira = real + 1j*imag
                #print(device_range)
                #mat[i][j] = dist
                #print(dist)
                #That plots the CIR contains in the buffer.
                #It still requires post-procesing to
                #re-align delay and received power level.
                #plt.plot((abs(cira[:-36])))
                dist_4 = device_range_4[1]-cal(cira,th)
                print(dist_4)
                #plt.show()
            except:
                print('Buffer Error on 4')
        else:
            print('error in getting cir')
    else:
        print('Ranging failed on 4')
        zd_4= device_range_4[1]

    # Multilateration Algorithm #
    import numpy as np

    def tri(zd, H):
        zd = zd.T
        H=H-H[0]
        tmp = zd[0]
        tmp = tmp * tmp
        zd = np.delete(zd,(0), axis=0)
        r1=np.multiply(zd,zd)
        Ha=np.delete(H,(0), axis=0)
        K=np.multiply(Ha,Ha)
        K=K.sum(axis=1)
        b= 0.5*(K-r1+tmp) 
        S_inv = np.linalg.inv(Ha.T*Ha)*Ha.T
        x_hat = S_inv*b
        return x_hat
    H = np.matrix([[7.54,0],[7.54,7.21],[14.14,0],[14.14,7.58]])
    zd = np.matrix([dist_1,dist_2,dist_3,dist_4])
    print(zd)
    x_hat = tri(zd*10**(-3),H)
    x_hat = x_hat.T+H[0]
    print(x_hat)
    zd_nlos = np.matrix([zd_1,zd_2,zd_3,zd_4])
    print(zd_nlos)
    x_nlos = tri(zd_nlos*10**(-3),H)
    x_nlos = x_nlos.T+H[0]
    print(x_nlos)
    write_ws['A1'] = 'NLOS'
    write_ws['C1'] = 'CIR'
    write_ws.cell(i+1,1,x_nlos[(0,0)])
    write_ws.cell(i+1,2,x_nlos[(0,1)])
    write_ws.cell(i+1,3,x_hat[(0,0)])
    write_ws.cell(i+1,4,x_hat[(0,1)])
    write_wb.save('/Users/swbaek/Desktop/CIR_EX_21.xlsx')
    i = i+1

